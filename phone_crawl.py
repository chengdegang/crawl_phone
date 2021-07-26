import os
import time
import bs4
import openpyxl
import requests
import re
from openpyxl import Workbook

"""
获取所有手机的品牌名称及相关链接
"""
def get_phonesinfo():
    url = 'https://github.com/KHwang9883/MobileModels'
    res = requests.get(url)
    res.raise_for_status()
    # 将页面源码写入本地
    with open('data/result.txt', 'wb') as result:
        for line in res.iter_content(200000):
            result.write(line)
    result.close()
    soup = bs4.BeautifulSoup(open('data/result.txt', 'r', encoding='utf-8'), 'html.parser')
    #遍历去取所有手机的链接、中英文名，19取的不严谨，后续获取该模块最大信息数，再更新
    urls = []
    for i in range(19):
        name_eng = soup.select(f'#readme > div.Box-body.px-5.pb-5 > article > table > tbody > tr:nth-child({i+1}) > td:nth-child(1) > a')
        name_cn = soup.select(f'#readme > div.Box-body.px-5.pb-5 > article > table > tbody > tr:nth-child({i+1}) > td:nth-child(2)')
        # print(name_eng)
        # print(name_cn)
        result_eng = ''.join(re.findall(r'">(.+?)</a>', str(name_eng)))
        url = ''.join(re.findall(r'href="(.+?)">', str(name_eng)))
        result_cn = ''.join(re.findall(r'>(.+?)</td>', str(name_cn)))
        print(f'{result_eng}、、https://github.com{url}、、{result_cn}')
        urls.append(url)
    # print(urls)
    return urls

"""
获取独立的每一个手机的信息
"""
def get_one_info(urls):
    #name1表示该品牌的model如iPhone 3G (iPhone1,2)，name2表示该model下细化的详情如A1324: iPhone 3G (国行)
    for i in range(len(urls)):
        res = requests.get(f'https://github.com{urls[i]}')
        soup = bs4.BeautifulSoup(res.text, 'html.parser')
        #获取当前页面所需信息
        # 获取title
        title = soup.select('#readme > article > h1')
        title = ''.join(re.findall(r'</a>(.+?)</h1>', str(title)))
        #遍历p下的所有数据，并分类
        name1 = []
        name2 = []
        for sibling in soup.p.next_siblings:  # 遍历后续节点
            if 'strong' in str(sibling):
                name2.append(str(sibling))
            else:
                name1.append(str(sibling))

        name1 = list(filter(lambda x: x != '\n', name1))
        name1 = list(filter(lambda x: x != '', name1))

        name1_detail1 = []
        # name1中细分name-1，-2
        for i in name1:
            name1_1 = re.findall(r'<p><code>(.+?)</code>', str(i))
            name1_2 = re.findall(r': (.+?)</p>', str(i))
            if name1_1 != '' and name1_2 != '':
                temp = name1_1 + name1_2
                name1_detail1.append(temp)
        # print(name1_detail1)
        name1_detail1 = [x for x in name1_detail1 if x]
        write_excel(file='data/Model summary.xlsx',data=name1_detail1,sheetname=title)

"""
输入的是一个以行为分隔的列表，若输入文件存在，则在该文件新建一个sheet存储数据，若不存在，则创建一个时间戳命名的xlsx文件
"""
def write_excel(file = '默认.xlsx',data = [['默认数据1'],['默认数据2']],sheetname = 'ces'):
    if os.path.exists(file) == True:
        wb = openpyxl.load_workbook(file)
        # 默认写到第一个sheet中，index为0
        wb.create_sheet(index=0, title=f'{sheetname}')
        sheet = wb[f'{sheetname}']
        # 好的写法写入excel
        for row_index, row_item in enumerate(data):
            for col_index, col_item in enumerate(row_item):
                sheet.cell(row=row_index + 1, column=col_index + 1, value=col_item)
        wb.save(file)
        print('---写入完成---')
    else:
        wbc = Workbook()
        log_path = os.path.dirname(os.path.abspath('.')) + '/testexcel/file/'
        t = time.strftime('%Y%m%d%H%M%S', time.localtime(time.time()))
        suffix = '.xlsx'  # 文件类型
        newfile = t + suffix
        path = log_path + t + suffix
        wbc.save(path)
        print(f"输入路径不存在xlsx文件,创建文件 {log_path+newfile} ")
        wb = openpyxl.load_workbook(log_path+newfile)
        # 默认写到第一个sheet中，index为0
        wb.create_sheet(index=0, title=f'{sheetname}')
        sheet = wb[f'{sheetname}']
        # 好的写法写入excel
        for row_index, row_item in enumerate(data):
            for col_index, col_item in enumerate(row_item):
                sheet.cell(row=row_index + 1, column=col_index + 1, value=col_item)
        wb.save(log_path+newfile)
        print('---写入完成---')

urls = ['/KHwang9883/MobileModels/blob/master/brands/apple.md']
if __name__ == '__main__':
    1
    #运行代码
    get_one_info(get_phonesinfo())
    print('运行完成')

